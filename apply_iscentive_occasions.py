import json
import os
from copy import copy
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
SOURCE = ROOT / "Новый Список ароматов FLUIDE (1).xlsx"
DATA = ROOT / "iscentive_occasions.json"
HEADER = "Случаи (сила 0–100)"

LABELS = {
    "everyday": "На каждый день",
    "evening": "Вечерние",
    "date": "Свидание",
    "gym": "Спортзал",
    "walk": "Прогулка",
}


def format_score(value):
    number = float(value)
    if number.is_integer():
        return str(int(number))
    return f"{number:.1f}".replace(".", ",")


def format_occasions(scores):
    canonical_order = {key: index for index, key in enumerate(LABELS)}
    ranked = sorted(
        LABELS,
        key=lambda key: (-float(scores.get(key, 0)), canonical_order[key]),
    )
    return "; ".join(
        f"{LABELS[key]}: {format_score(scores.get(key, 0))}"
        for key in ranked
    )


payload = json.loads(DATA.read_text(encoding="utf-8"))
items = payload["items"]

workbook = load_workbook(SOURCE)
sheet = workbook.active

headers = {
    str(cell.value).strip(): cell.column
    for cell in sheet[1]
    if cell.value is not None
}
column = headers.get(HEADER, sheet.max_column + 1)

template_header = sheet.cell(row=1, column=column - 1)
header_cell = sheet.cell(row=1, column=column, value=HEADER)
header_cell._style = copy(template_header._style)
header_cell.alignment = copy(template_header.alignment)
header_cell.number_format = template_header.number_format
header_cell.protection = copy(template_header.protection)

updated = 0
for row in range(2, sheet.max_row + 1):
    raw_number = sheet.cell(row=row, column=1).value
    if raw_number is None:
        continue
    number = str(raw_number).strip().zfill(3)
    item = items.get(number)
    if not item:
        raise KeyError(f"Нет данных случаев для аромата {number}")

    template_cell = sheet.cell(row=row, column=column - 1)
    target = sheet.cell(
        row=row,
        column=column,
        value=format_occasions(item["occasionScores"]),
    )
    target._style = copy(template_cell._style)
    target.alignment = copy(template_cell.alignment)
    target.number_format = template_cell.number_format
    target.protection = copy(template_cell.protection)
    updated += 1

previous_letter = sheet.cell(row=1, column=column - 1).column_letter
target_letter = sheet.cell(row=1, column=column).column_letter
sheet.column_dimensions[target_letter].width = max(
    58,
    sheet.column_dimensions[previous_letter].width or 0,
)
sheet.auto_filter.ref = f"A1:{target_letter}{sheet.max_row}"

temporary = SOURCE.with_name(f"{SOURCE.stem}.tmp{SOURCE.suffix}")
workbook.save(temporary)
os.replace(temporary, SOURCE)

print(f"Обновлён {SOURCE.name}: {updated} ароматов, столбец {target_letter}")
