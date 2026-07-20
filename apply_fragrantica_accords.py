import json
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent
WORKBOOK_PATH = ROOT / "Новый Список ароматов FLUIDE (1).xlsx"
ACCORDS_PATH = ROOT / "fragrantica_accords.json"
HEADER = "Аккорды (сила 0–100)"
MISSING_LABEL = "Нет страницы на Fragrantica"


def format_weight(value):
    number = float(value)
    if number.is_integer():
        return str(int(number))
    return f"{number:.1f}".replace(".", ",")


def format_accords(item):
    accords = item.get("accords", [])
    if not accords:
        return MISSING_LABEL if item.get("status") == "source_page_not_found" else ""
    return "; ".join(
        f"{accord['name']}: {format_weight(accord['weight'])}"
        for accord in accords
    )


data = json.loads(ACCORDS_PATH.read_text(encoding="utf-8"))
items = data["items"]
workbook = load_workbook(WORKBOOK_PATH)
sheet = workbook.active

headers = {
    str(cell.value).strip(): cell.column
    for cell in sheet[1]
    if cell.value is not None
}
target_column = headers.get(HEADER, sheet.max_column + 1)
source_style_column = max(1, target_column - 1)

header_source = sheet.cell(row=1, column=source_style_column)
header_target = sheet.cell(row=1, column=target_column, value=HEADER)
header_target._style = copy(header_source._style)
header_target.font = copy(header_source.font)
header_target.fill = copy(header_source.fill)
header_target.border = copy(header_source.border)
header_target.protection = copy(header_source.protection)
header_target.alignment = Alignment(
    horizontal=header_source.alignment.horizontal,
    vertical=header_source.alignment.vertical or "center",
    text_rotation=header_source.alignment.text_rotation,
    wrap_text=True,
    shrink_to_fit=header_source.alignment.shrink_to_fit,
    indent=header_source.alignment.indent,
)

written = 0
missing = []
for row_number in range(2, sheet.max_row + 1):
    raw_number = sheet.cell(row=row_number, column=1).value
    if raw_number in (None, ""):
        continue
    number = str(raw_number).strip().zfill(3)
    item = items.get(number)
    if item is None:
        missing.append(number)
        continue

    source_cell = sheet.cell(row=row_number, column=source_style_column)
    target_cell = sheet.cell(
        row=row_number,
        column=target_column,
        value=format_accords(item),
    )
    target_cell._style = copy(source_cell._style)
    target_cell.font = copy(source_cell.font)
    target_cell.fill = copy(source_cell.fill)
    target_cell.border = copy(source_cell.border)
    target_cell.protection = copy(source_cell.protection)
    target_cell.number_format = source_cell.number_format
    target_cell.alignment = Alignment(
        horizontal="left",
        vertical="top",
        wrap_text=True,
    )
    written += 1

column_letter = get_column_letter(target_column)
sheet.column_dimensions[column_letter].width = 58
if sheet.auto_filter.ref:
    start, end = sheet.auto_filter.ref.split(":")
    end_row = "".join(character for character in end if character.isdigit())
    sheet.auto_filter.ref = f"{start}:{column_letter}{end_row}"

temporary_path = WORKBOOK_PATH.with_name(f"{WORKBOOK_PATH.stem}.tmp.xlsx")
workbook.save(temporary_path)
temporary_path.replace(WORKBOOK_PATH)

if missing:
    raise RuntimeError(f"Для номеров не найдены данные: {', '.join(missing)}")

print(
    f"Обновлена таблица: {written} строк, "
    f"столбец {column_letter} «{HEADER}»"
)
