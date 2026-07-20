import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook


sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parent
SOURCE = Path(sys.argv[1]).resolve() if len(sys.argv) > 1 else ROOT / "Новый Список ароматов FLUIDE (1).xlsx"
OUTPUT = ROOT / "fragrances.json"
IMAGES = ROOT / "images" / "fragrances"

FAMILY_KEYWORDS = {
    "Цветочные": [
        "роза", "жасмин", "пион", "ирис", "фиал", "ландыш", "тубероз",
        "магноли", "орхиде", "лаванд", "цветок", "османтус", "герань", "нероли",
    ],
    "Фруктовые": [
        "яблок", "груш", "персик", "абрикос", "слив", "виш", "череш", "ананас",
        "манго", "маракуй", "ягод", "смород", "малин", "клубник", "гранат", "дын",
        "арбуз", "инжир", "личи",
    ],
    "Цитрусовые": [
        "бергамот", "лимон", "мандарин", "апельсин", "грейпфрут", "лайм",
        "цитрус", "юдзу", "помело",
    ],
    "Древесные": [
        "кедр", "сандал", "древес", "ветивер", "пачули", "уд", "мох", "кипарис", "гуаяк",
    ],
    "Сладкие": [
        "ваниль", "карамел", "шоколад", "какао", "мед", "пралине", "сахар", "зефир",
        "бобы тонка", "корица",
    ],
    "Свежие": [
        "морск", "водн", "озон", "акват", "свеж", "зелен", "мята", "лед", "огур", "чай",
    ],
    "Пряные и восточные": [
        "перец", "имбир", "кардамон", "шафран", "гвоздик", "мускат", "табак", "кожа",
        "ладан", "амбра", "мускус",
    ],
}

FAMILY_VALUES = [
    "Цветочные",
    "Фруктовые",
    "Цитрусовые",
    "Древесные",
    "Сладкие",
    "Свежие",
    "Пряные и восточные",
]

GROUP_FAMILY_KEYWORDS = {
    "Цветочные": ["цветоч"],
    "Фруктовые": ["фруктов"],
    "Цитрусовые": ["цитрусов"],
    "Древесные": ["древесн", "шипров"],
    "Сладкие": ["гурманск", "сладк"],
    "Свежие": ["водян", "зелен", "фужерн"],
    "Пряные и восточные": ["восточн", "прян", "кожан"],
}

SEASON_LABELS = {
    "лето": "summer",
    "осень": "autumn",
    "зима": "winter",
    "весна": "spring",
}

OCCASION_LABELS = {
    "на каждый день": "everyday",
    "вечерние": "evening",
    "свидание": "date",
    "спортзал": "gym",
    "прогулка": "walk",
}

OCCASION_VALUES = ["everyday", "evening", "date", "gym", "walk"]
SEASON_VALUES = ["summer", "autumn", "winter", "spring"]

ACCORD_FAMILY_MAP = {
    "Цветочные": {
        "цветочный", "белые цветы", "розовый", "желтые цветы", "тубероза",
        "фиалковый", "ирис",
    },
    "Фруктовые": {
        "фруктовый", "тропический", "вишневый", "кокосовый",
    },
    "Цитрусовые": {"цитрусовый"},
    "Древесные": {
        "древесный", "пачулиевый", "землистый", "мшистый", "удовый", "хвойный",
    },
    "Сладкие": {
        "сладкий", "ванильный", "карамельный", "какао", "ореховый", "кофейный",
        "шоколад", "медовый", "миндальный", "лактонный",
    },
    "Свежие": {
        "свежий", "фужерный", "зеленый", "акватический", "морской", "озоновый",
        "лаванда", "травяной", "соленый", "минеральный", "мыльный", "альдегидный",
        "металлический", "камфорный", "каннабис", "водочный",
    },
    "Пряные и восточные": {
        "свежий пряный", "теплый пряный", "мягкий пряный", "амбровый", "мускусный",
        "животный", "бальзамический", "кожаный", "табачный", "дымный", "коричный",
        "удовый", "пудровый",
    },
}

FRESH_KEYWORDS = [
    "морск", "водн", "озон", "акват", "свеж", "зелен", "мята", "чай",
    "бергамот", "лимон", "мандарин", "апельсин", "грейпфрут", "лайм", "юдзу",
    "базилик", "лаванд", "нероли", "шалфей", "соль", "имбир",
]
SOFT_FLORAL_KEYWORDS = [
    "роза", "жасмин", "пион", "ирис", "фиал", "ландыш", "магноли",
    "орхиде", "гардени", "тубероз", "османтус", "фрез", "цвет",
]
JUICY_KEYWORDS = [
    "яблок", "груш", "персик", "абрикос", "слив", "виш", "череш", "ананас",
    "манго", "маракуй", "ягод", "смород", "малин", "клубник", "гранат",
    "дын", "арбуз", "инжир", "личи", "кокос",
]
WARM_KEYWORDS = [
    "ваниль", "карамел", "шоколад", "какао", "мед", "пралине", "сахар",
    "зефир", "бобы тонка", "корица", "амбра", "табак", "кожа",
    "ладан", "уд", "шафран", "кардамон", "гвоздик", "ром",
    "коньяк", "кофе", "пачули", "сандал",
]
DEEP_KEYWORDS = [
    "табак", "кожа", "ладан", "уд", "дым", "каннабис", "абсент", "кофе",
    "шоколад", "какао", "ром", "коньяк", "пачули", "амбра",
]

OCCASION_OVERRIDES = {
    "003": ["everyday", "gym", "walk"],
    "018": ["everyday", "gym", "walk"],
    "022": ["everyday", "gym", "walk"],
    "023": ["everyday", "walk"],
    "030": ["everyday", "date", "walk"],
    "031": ["everyday", "gym", "walk"],
    "036": ["everyday", "gym", "walk"],
    "050": ["evening", "date"],
    "070": ["everyday", "gym", "walk"],
    "071": ["everyday", "gym", "walk"],
    "090": ["everyday", "gym", "walk"],
    "148": ["everyday", "gym", "walk"],
    "156": ["everyday", "gym", "walk"],
    "183": ["everyday", "gym", "walk"],
    "195": ["everyday", "gym", "walk"],
    "198": ["everyday", "gym", "walk"],
    "505": ["everyday", "gym", "walk"],
    "517": ["everyday", "gym", "walk"],
}

SEASON_OVERRIDES = {
    "003": ["summer", "spring"],
    "018": ["summer", "spring"],
    "020": ["summer"],
    "022": ["summer", "spring"],
    "030": ["summer", "spring"],
    "031": ["summer", "spring"],
    "036": ["summer", "spring"],
    "041": ["autumn", "winter"],
    "042": ["autumn", "winter"],
    "043": ["autumn", "winter"],
    "050": ["autumn", "winter"],
    "070": ["summer", "spring"],
    "071": ["summer", "spring"],
    "090": ["summer", "spring"],
    "148": ["summer", "spring"],
    "170": ["autumn", "winter"],
    "183": ["summer", "spring"],
    "195": ["summer"],
    "198": ["summer", "spring"],
    "505": ["summer", "spring"],
    "517": ["summer", "spring"],
    "526": ["autumn", "winter"],
}

GENDER_OVERRIDES = {
    # В исходной таблице Eclat d'Arpege ошибочно указан как мужской.
    "030": "женский",
}


def split_notes(raw):
    groups = {"top": [], "middle": [], "base": [], "main": []}
    if not raw:
        return groups

    labels = {
        "верхние": "top",
        "средние": "middle",
        "базовые": "base",
        "основные": "main",
    }
    for line in str(raw).replace("\r", "").split("\n"):
        if ":" not in line:
            continue
        label, values = line.split(":", 1)
        key = labels.get(label.strip().lower())
        if not key:
            continue
        groups[key] = [item.strip() for item in values.split(",") if item.strip()]
    return groups


def detect_families(raw):
    text = str(raw or "").lower()
    families = [
        family
        for family, keywords in FAMILY_KEYWORDS.items()
        if any(keyword in text for keyword in keywords)
    ]
    return families or ["Другие"]


def detect_group_families(group):
    text = str(group or "").strip().lower()
    return [
        family
        for family, keywords in GROUP_FAMILY_KEYWORDS.items()
        if any(keyword in text for keyword in keywords)
    ]


def parse_seasons(raw):
    values = []
    for value in str(raw or "").split(","):
        normalized = SEASON_LABELS.get(value.strip().lower())
        if normalized:
            values.append(normalized)
    return ordered_values(values, SEASON_VALUES)


def parse_accords(raw):
    accords = []
    for item in re.split(r"[;\n]+", str(raw or "")):
        item = item.strip()
        if not item or ":" not in item:
            continue
        name, weight_raw = item.rsplit(":", 1)
        try:
            weight = float(weight_raw.strip().replace(",", "."))
        except ValueError:
            continue
        accords.append(
            {
                "name": name.strip(),
                "weight": int(weight) if weight.is_integer() else weight,
            }
        )
    return accords


def parse_occasion_scores(raw):
    scores = {}
    for item in re.split(r"[;\n]+", str(raw or "")):
        item = item.strip()
        if not item or ":" not in item:
            continue
        label, score_raw = item.rsplit(":", 1)
        key = OCCASION_LABELS.get(label.strip().lower())
        if not key:
            continue
        try:
            score = float(score_raw.strip().replace(",", "."))
        except ValueError:
            continue
        scores[key] = int(score) if score.is_integer() else score
    return scores


def build_family_scores(accords, group_families, note_families):
    accord_totals = {}
    for family in FAMILY_VALUES:
        weights = sorted(
            (
                float(accord["weight"])
                for accord in accords
                if str(accord["name"]).strip().lower() in ACCORD_FAMILY_MAP[family]
            ),
            reverse=True,
        )
        # Ведущий аккорд определяет направление, следующие уточняют сложность,
        # но не могут победить только за счёт большого количества слабых полос.
        coefficients = (1.0, 0.35, 0.15)
        accord_totals[family] = sum(
            weight * coefficients[index]
            for index, weight in enumerate(weights[: len(coefficients)])
        )

    peak = max(accord_totals.values(), default=0)
    accord_scores = {
        family: (total / peak * 100 if peak else 0)
        for family, total in accord_totals.items()
    }

    combined = {}
    for family in FAMILY_VALUES:
        accord_score = accord_scores[family]
        group_score = 100 if family in group_families else 0
        note_fallback = 30 if family in note_families else 0
        if accords:
            value = 0.65 * accord_score + 0.35 * group_score
        else:
            value = max(group_score, note_fallback)
        combined[family] = value

    combined_peak = max(combined.values(), default=0)
    if not combined_peak:
        return {family: 0 for family in FAMILY_VALUES}
    return {
        family: round(value / combined_peak * 100, 1)
        for family, value in combined.items()
    }


def select_occasions(scores):
    if not scores:
        return []
    selected = [key for key, score in scores.items() if float(score) >= 60]
    if not selected:
        selected = [max(scores, key=lambda key: float(scores[key]))]
    return ordered_values(selected, OCCASION_VALUES)


def has_any(text, keywords):
    return any(keyword in text for keyword in keywords)


def count_keywords(text, keywords):
    return sum(1 for keyword in keywords if keyword in text)


def ordered_values(values, order):
    unique = set(values)
    return [value for value in order if value in unique]


def detect_occasions(number, notes_raw, families, oil_percent, title, original):
    if number in OCCASION_OVERRIDES:
        return OCCASION_OVERRIDES[number]

    text = f"{notes_raw} {title} {original}".lower()
    fresh_score = count_keywords(text, FRESH_KEYWORDS)
    floral_score = count_keywords(text, SOFT_FLORAL_KEYWORDS)
    juicy_score = count_keywords(text, JUICY_KEYWORDS)
    warm_score = count_keywords(text, WARM_KEYWORDS)
    deep_score = count_keywords(text, DEEP_KEYWORDS)
    oil = oil_percent or 25

    occasions = []
    if (oil <= 28 and deep_score == 0) or fresh_score >= 2:
        occasions.append("everyday")
    if fresh_score >= 3 and oil <= 28 and warm_score <= 1 and deep_score == 0:
        occasions.append("gym")
    if fresh_score + floral_score + juicy_score >= 3 and oil <= 30 and deep_score == 0:
        occasions.append("walk")
    if floral_score >= 2 or juicy_score >= 2 or (warm_score >= 2 and deep_score == 0):
        occasions.append("date")
    if oil >= 30 or warm_score >= 3 or deep_score > 0:
        occasions.append("evening")
    if deep_score > 0 and "gym" in occasions:
        occasions.remove("gym")
    if not occasions:
        occasions.append("everyday")
    return ordered_values(occasions, OCCASION_VALUES)


def detect_seasons(number, notes_raw, families, oil_percent, title, original):
    if number in SEASON_OVERRIDES:
        return SEASON_OVERRIDES[number]

    text = f"{notes_raw} {title} {original}".lower()
    fresh_score = count_keywords(text, FRESH_KEYWORDS)
    floral_score = count_keywords(text, SOFT_FLORAL_KEYWORDS)
    juicy_score = count_keywords(text, JUICY_KEYWORDS)
    warm_score = count_keywords(text, WARM_KEYWORDS)
    deep_score = count_keywords(text, DEEP_KEYWORDS)
    oil = oil_percent or 25

    seasons = []
    if fresh_score >= 2 and warm_score <= 2 and deep_score == 0:
        seasons.append("summer")
    if fresh_score + floral_score + juicy_score >= 3 and deep_score == 0:
        seasons.append("spring")
    if warm_score >= 2 or (juicy_score >= 2 and oil >= 25) or ("Древесные" in families and oil >= 25):
        seasons.append("autumn")
    if deep_score > 0 or warm_score >= 4 or (oil >= 30 and warm_score >= 2):
        seasons.append("winter")
    if not seasons:
        seasons.extend(["spring", "autumn"])
    return ordered_values(seasons, SEASON_VALUES)


def clean_title(value):
    title = re.sub(r"^FLUIDE\s+", "", str(value or ""), flags=re.IGNORECASE)
    title = re.sub(r"\s+30\s*мл$", "", title, flags=re.IGNORECASE)
    return title.strip()


workbook = load_workbook(SOURCE, read_only=True, data_only=True)
sheet = workbook.active
fragrances = []

for row in sheet.iter_rows(min_row=2, values_only=True):
    if not row[0]:
        continue
    number = str(row[0]).strip().zfill(3)
    notes_raw = str(row[9] or "").strip()
    oil_percent = int(row[3]) if row[3] is not None else None
    title = clean_title(row[1])
    original = str(row[2] or "").strip()
    group = str(row[10] or "").strip() if len(row) > 10 else ""
    group_families = detect_group_families(group)
    note_families = detect_families(notes_raw)
    families = ordered_values(group_families + note_families, FAMILY_VALUES)
    seasons = parse_seasons(row[11] if len(row) > 11 else "")
    accords = parse_accords(row[12] if len(row) > 12 else "")
    family_scores = build_family_scores(accords, group_families, note_families)
    occasion_scores = parse_occasion_scores(row[13] if len(row) > 13 else "")
    occasions = select_occasions(occasion_scores)
    if not occasions:
        occasions = detect_occasions(number, notes_raw, families, oil_percent, title, original)
    if not seasons:
        seasons = detect_seasons(number, notes_raw, families, oil_percent, title, original)
    image_path = IMAGES / f"{number}.webp"
    thumbnail_path = IMAGES / "thumbs" / f"{number}.webp"
    fragrances.append(
        {
            "id": number,
            "name": str(row[1] or "").strip(),
            "title": title,
            "original": original,
            "oilPercent": oil_percent,
            "gender": GENDER_OVERRIDES.get(number, str(row[5] or "").strip().lower()),
            "category": str(row[6] or "").strip(),
            "concentration": str(row[8] or "").strip(),
            "notes": split_notes(notes_raw),
            "notesRaw": notes_raw,
            "group": group,
            "groupFamilies": group_families,
            "families": families,
            "familyScores": family_scores,
            "occasion": occasions,
            "occasionScores": occasion_scores,
            "season": seasons,
            "accords": accords,
            **({"image": f"images/fragrances/{number}.webp"} if image_path.exists() else {}),
            **({"thumbnail": f"images/fragrances/thumbs/{number}.webp"} if thumbnail_path.exists() else {}),
        }
    )

OUTPUT.write_text(json.dumps(fragrances, ensure_ascii=False, indent=2), encoding="utf-8")
print(f"Создан {OUTPUT.name}: {len(fragrances)} ароматов")
